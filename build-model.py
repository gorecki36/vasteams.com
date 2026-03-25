#!/usr/bin/env python3
"""
Build the Matrix Outcome Model spreadsheet.
All formulas are live Excel formulas — no hardcoded numbers.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from copy import copy

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────

DARK_BG = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
HEADER_BG = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
INPUT_BG = PatternFill(start_color="1A2E1A", end_color="1A2E1A", fill_type="solid")
OUTCOME_BG = PatternFill(start_color="2E1A1A", end_color="2E1A1A", fill_type="solid")
MATRIX_BG = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
PARAM_BG = PatternFill(start_color="2E2E1A", end_color="2E2E1A", fill_type="solid")

WHITE = Font(color="FFFFFF", size=11)
WHITE_BOLD = Font(color="FFFFFF", size=11, bold=True)
GREEN_FONT = Font(color="34D399", size=11, bold=True)
RED_FONT = Font(color="F87171", size=11, bold=True)
AMBER_FONT = Font(color="FBBF24", size=11, bold=True)
BLUE_FONT = Font(color="60A5FA", size=11, bold=True)
TITLE_FONT = Font(color="34D399", size=14, bold=True)
SECTION_FONT = Font(color="FBBF24", size=12, bold=True)
SMALL_WHITE = Font(color="A1A1AA", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="3F3F46"),
    right=Side(style="thin", color="3F3F46"),
    top=Side(style="thin", color="3F3F46"),
    bottom=Side(style="thin", color="3F3F46"),
)

def style_range(ws, row, col_start, col_end, font=WHITE, fill=DARK_BG, border=True):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        if border:
            cell.border = THIN_BORDER

def style_cell(ws, row, col, font=WHITE, fill=DARK_BG, border=True):
    cell = ws.cell(row=row, column=col)
    cell.font = font
    cell.fill = fill
    if border:
        cell.border = THIN_BORDER
    return cell


# ═══════════════════════════════════════════════════════════════
# SHEET 1: INPUTS
# ═══════════════════════════════════════════════════════════════

ws1 = wb.active
ws1.title = "Inputs"
ws1.sheet_properties.tabColor = "34D399"

# Dark background for all cells
for row in range(1, 50):
    for col in range(1, 12):
        ws1.cell(row=row, column=col).fill = DARK_BG

# Column widths
ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 25
ws1.column_dimensions["C"].width = 15
ws1.column_dimensions["D"].width = 15
ws1.column_dimensions["E"].width = 40

# Title
ws1.merge_cells("B2:E2")
style_cell(ws1, 2, 2, TITLE_FONT)
ws1["B2"] = "MATRIX OUTCOME MODEL"

ws1.merge_cells("B3:E3")
style_cell(ws1, 3, 2, SMALL_WHITE)
ws1["B3"] = "Adjust the 7 input sliders. Everything else is computed."

# ── 7 Input Forces ──────────────────────────────────────────

row = 5
style_cell(ws1, row, 2, SECTION_FONT)
ws1[f"B{row}"] = "INPUT FORCES"
style_cell(ws1, row, 3, SECTION_FONT)
ws1[f"C{row}"] = "Value (0-100)"
style_cell(ws1, row, 4, SECTION_FONT)
ws1[f"D{row}"] = "Default"
style_cell(ws1, row, 5, SECTION_FONT)
ws1[f"E{row}"] = "Description"

inputs = [
    ("ai_acceleration", "AI Acceleration", 50, "Rate of AI capability growth"),
    ("climate_crisis", "Climate Crisis", 50, "Severity of climate disruption"),
    ("neural_interface", "Neural Interface Tech", 50, "Brain-computer interface progress"),
    ("corporate_power", "Corporate Power", 50, "Concentration of corporate control"),
    ("demographic_collapse", "Demographic Collapse", 50, "Rate of population decline"),
    ("digital_immersion", "Digital Immersion", 50, "Share of life in digital environments"),
    ("global_convergence", "Global Convergence", 50, "Whether regions converge or diverge"),
]

INPUT_START_ROW = 6
for i, (id_, name, default, desc) in enumerate(inputs):
    r = INPUT_START_ROW + i
    style_cell(ws1, r, 2, WHITE_BOLD, INPUT_BG)
    ws1[f"B{r}"] = name
    cell = style_cell(ws1, r, 3, GREEN_FONT, INPUT_BG)
    cell.value = default
    cell.number_format = "0"
    style_cell(ws1, r, 4, SMALL_WHITE, INPUT_BG)
    ws1[f"D{r}"] = default
    style_cell(ws1, r, 5, SMALL_WHITE, INPUT_BG)
    ws1[f"E{r}"] = desc

# Named references for readability
# C6=AI, C7=Climate, C8=Neural, C9=Corporate, C10=Demo, C11=Digital, C12=Global
INPUT_END_ROW = INPUT_START_ROW + len(inputs) - 1

# ── Model Parameters ────────────────────────────────────────

row = INPUT_END_ROW + 2  # row 14
style_cell(ws1, row, 2, SECTION_FONT)
ws1[f"B{row}"] = "MODEL PARAMETERS"

params = [
    ("k (steepness)", 0.065, "Controls transition speed — ~68yr from 10% to 90%"),
    ("Latest arrival year", 2125, "Arrival year when pressure = 0"),
    ("Arrival range (years)", 70, "How many years earlier when pressure = 100"),
    ("Tier 2 lag (years)", 5, "Extra delay before downstream outcomes kick in"),
]

PARAM_START_ROW = row + 1
for i, (name, val, desc) in enumerate(params):
    r = PARAM_START_ROW + i
    style_cell(ws1, r, 2, WHITE_BOLD, PARAM_BG)
    ws1[f"B{r}"] = name
    cell = style_cell(ws1, r, 3, BLUE_FONT, PARAM_BG)
    cell.value = val
    if isinstance(val, float):
        cell.number_format = "0.000"
    else:
        cell.number_format = "0"
    style_cell(ws1, r, 5, SMALL_WHITE, PARAM_BG)
    ws1[f"E{r}"] = desc

# Cell references for params
K_CELL = f"C{PARAM_START_ROW}"        # C15
LATEST_CELL = f"C{PARAM_START_ROW+1}" # C16
RANGE_CELL = f"C{PARAM_START_ROW+2}"  # C17
LAG_CELL = f"C{PARAM_START_ROW+3}"    # C18

# ── How to read ─────────────────────────────────────────────

row = PARAM_START_ROW + len(params) + 1  # row 20
style_cell(ws1, row, 2, SECTION_FONT)
ws1[f"B{row}"] = "HOW IT WORKS"
notes = [
    "1. Each outcome has a 'pressure' (0-100) = weighted sum of inputs",
    "2. Pressure determines the 'arrival year' — when that outcome crosses 50%",
    "3. Higher pressure = earlier arrival = the crisis comes sooner",
    "4. The outcome over time is a sigmoid: always 0→100, just shifted",
    "5. Formula: outcome(year) = 100 / (1 + EXP(-k × (year - arrival_year)))",
    "6. Moment of Matrix = weighted average of all 5 outcomes",
    "",
    "→ Go to the 'Outcomes' tab to see the weight matrix",
    "→ Go to the 'Timeline' tab to see the projections",
]
for i, note in enumerate(notes):
    r = row + 1 + i
    style_cell(ws1, r, 2, SMALL_WHITE)
    ws1.merge_cells(f"B{r}:E{r}")
    ws1[f"B{r}"] = note


# ═══════════════════════════════════════════════════════════════
# SHEET 2: OUTCOMES (weight matrix + pressure + arrival)
# ═══════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Outcomes")
ws2.sheet_properties.tabColor = "F87171"

for row in range(1, 40):
    for col in range(1, 15):
        ws2.cell(row=row, column=col).fill = DARK_BG

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 25
ws2.column_dimensions["C"].width = 16  # AI Accel
ws2.column_dimensions["D"].width = 14  # Climate
ws2.column_dimensions["E"].width = 16  # Neural
ws2.column_dimensions["F"].width = 16  # Corporate
ws2.column_dimensions["G"].width = 18  # Demographic
ws2.column_dimensions["H"].width = 16  # Digital
ws2.column_dimensions["I"].width = 16  # Global
ws2.column_dimensions["J"].width = 4   # spacer
ws2.column_dimensions["K"].width = 16  # Pressure
ws2.column_dimensions["L"].width = 16  # Arrival Year

# Title
ws2.merge_cells("B2:L2")
style_cell(ws2, 2, 2, TITLE_FONT)
ws2["B2"] = "OUTCOME WEIGHT MATRIX"

ws2.merge_cells("B3:L3")
style_cell(ws2, 3, 2, SMALL_WHITE)
ws2["B3"] = "Weights show how much each input contributes to each outcome. Pressure & arrival are computed."

# Header row
row = 5
headers = ["Outcome", "AI Accel", "Climate", "Neural", "Corporate", "Demographic", "Digital", "Global", "", "Pressure", "Arrival Year"]
for i, h in enumerate(headers):
    c = 2 + i
    style_cell(ws2, row, c, WHITE_BOLD, HEADER_BG)
    ws2.cell(row=row, column=c).value = h

# Input values reference row
row = 6
style_cell(ws2, row, 2, SMALL_WHITE, HEADER_BG)
ws2[f"B{row}"] = "(input values →)"
input_cols = ["C", "D", "E", "F", "G", "H", "I"]
input_rows_ref = [6, 7, 8, 9, 10, 11, 12]  # rows in Inputs sheet
for i, col in enumerate(input_cols):
    cell = style_cell(ws2, row, 3 + i, GREEN_FONT, HEADER_BG)
    cell.value = f"=Inputs!C{input_rows_ref[i]}"
    cell.number_format = "0"

# ── Weight matrix ───────────────────────────────────────────

# Columns: B=name, C=AI, D=Climate, E=Neural, F=Corporate, G=Demo, H=Digital, I=Global, J=spacer, K=Pressure, L=Arrival

# Tier 1 outcomes
outcomes = [
    # (name, AI, Climate, Neural, Corporate, Demo, Digital, Global, tier)
    ("Meaning Crisis",     0.25, 0,    0,    0.20, 0.20, 0.35, 0,    1),
    ("Trust Collapse",     0.25, 0.25, 0,    0.30, 0,    0.20, 0,    1),
    ("Pod Economics",      0.25, 0,    0.30, 0.25, 0,    0,    0.20, 1),
    ("Escapism / Spiritual", 0, 0.20, 0.20, 0,    0,    0.25, 0,    2),  # + meaning_pressure*0.35
    ("Economic Withdrawal", 0.35, 0,   0,    0.25, 0,    0,    0,    2),  # + meaning_pressure*0.40
]

OUTCOME_START_ROW = 8
for i, (name, ai, clim, neur, corp, demo, digi, glob, tier) in enumerate(outcomes):
    r = OUTCOME_START_ROW + i
    bg = OUTCOME_BG
    style_cell(ws2, r, 2, WHITE_BOLD if tier == 1 else RED_FONT, bg)
    ws2[f"B{r}"] = name + (" (Tier 2)" if tier == 2 else "")

    weights = [ai, clim, neur, corp, demo, digi, glob]
    for j, w in enumerate(weights):
        cell = style_cell(ws2, r, 3 + j, WHITE if w > 0 else SMALL_WHITE, bg)
        cell.value = w
        cell.number_format = "0.00"

    # Pressure formula (column K = 11)
    # For Tier 1: SUMPRODUCT of weights × input values
    if tier == 1:
        # K = C_weight*C_input + D_weight*D_input + ...
        parts = []
        for j in range(7):
            w_cell = f"{get_column_letter(3+j)}{r}"
            v_cell = f"{get_column_letter(3+j)}6"  # input value reference row
            parts.append(f"{w_cell}*{v_cell}")
        formula = "=" + "+".join(parts)
    else:
        # Tier 2: includes meaning_pressure reference
        meaning_pressure_cell = f"K{OUTCOME_START_ROW}"  # K8 = meaning crisis pressure
        if "Escapism" in name:
            # meaning_pressure*0.35 + digital*0.25 + climate*0.20 + neural*0.20
            parts = [f"{meaning_pressure_cell}*0.35"]
            for j, w in enumerate(weights):
                if w > 0:
                    v_cell = f"{get_column_letter(3+j)}6"
                    parts.append(f"{w}*{v_cell}")
            formula = "=" + "+".join(parts)
        else:  # Economic Withdrawal
            # ai*0.35 + meaning_pressure*0.40 + corporate*0.25
            parts = [
                f"C6*0.35",  # AI
                f"{meaning_pressure_cell}*0.40",
                f"F6*0.25",  # Corporate
            ]
            formula = "=" + "+".join(parts)

    cell = style_cell(ws2, r, 11, BLUE_FONT, bg)
    cell.value = formula
    cell.number_format = "0.0"

    # Arrival year formula (column L = 12)
    # = LATEST - (pressure/100) * RANGE [+ LAG for Tier 2]
    if tier == 1:
        arr_formula = f"=Inputs!{LATEST_CELL}-(K{r}/100)*Inputs!{RANGE_CELL}"
    else:
        arr_formula = f"=Inputs!{LATEST_CELL}-(K{r}/100)*Inputs!{RANGE_CELL}+Inputs!{LAG_CELL}"
    cell = style_cell(ws2, r, 12, AMBER_FONT, bg)
    cell.value = arr_formula
    cell.number_format = "0"

OUTCOME_END_ROW = OUTCOME_START_ROW + len(outcomes) - 1  # row 12

# ── Moment of Matrix weights ───────────────────────────────

row = OUTCOME_END_ROW + 2  # row 14
style_cell(ws2, row, 2, SECTION_FONT)
ws2[f"B{row}"] = "MOMENT OF MATRIX"
style_cell(ws2, row, 11, SECTION_FONT)
ws2[f"K{row}"] = "Weight"

mom_weights = [
    ("Meaning Crisis", 0.20, OUTCOME_START_ROW),
    ("Trust Collapse", 0.15, OUTCOME_START_ROW + 1),
    ("Pod Economics", 0.15, OUTCOME_START_ROW + 2),
    ("Escapism / Spiritual", 0.25, OUTCOME_START_ROW + 3),
    ("Economic Withdrawal", 0.25, OUTCOME_START_ROW + 4),
]

MOM_START_ROW = row + 1
for i, (name, weight, _) in enumerate(mom_weights):
    r = MOM_START_ROW + i
    style_cell(ws2, r, 2, WHITE_BOLD, MATRIX_BG)
    ws2[f"B{r}"] = name
    cell = style_cell(ws2, r, 11, BLUE_FONT, MATRIX_BG)
    cell.value = weight
    cell.number_format = "0.00"


# ═══════════════════════════════════════════════════════════════
# SHEET 3: TIMELINE (year × outcome matrix with live formulas)
# ═══════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("Timeline")
ws3.sheet_properties.tabColor = "60A5FA"

for row in range(1, 50):
    for col in range(1, 12):
        ws3.cell(row=row, column=col).fill = DARK_BG

ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 10  # Year
ws3.column_dimensions["C"].width = 16  # Meaning
ws3.column_dimensions["D"].width = 16  # Trust
ws3.column_dimensions["E"].width = 16  # Pod
ws3.column_dimensions["F"].width = 18  # Escapism
ws3.column_dimensions["G"].width = 18  # Withdrawal
ws3.column_dimensions["H"].width = 4   # spacer
ws3.column_dimensions["I"].width = 18  # Matrix

# Title
ws3.merge_cells("B2:I2")
style_cell(ws3, 2, 2, TITLE_FONT)
ws3["B2"] = "TIMELINE PROJECTIONS"

ws3.merge_cells("B3:I3")
style_cell(ws3, 3, 2, SMALL_WHITE)
ws3["B3"] = "All values are live formulas. Change inputs on the Inputs tab and watch these update."

# Sigmoid formula explanation
ws3.merge_cells("B4:I4")
style_cell(ws3, 4, 2, SMALL_WHITE)
ws3["B4"] = "Formula: 100 / (1 + EXP(-k × (year - arrival_year)))"

# Headers
row = 6
headers = ["Year", "Meaning Crisis", "Trust Collapse", "Pod Economics", "Escapism", "Withdrawal", "", "MATRIX"]
for i, h in enumerate(headers):
    c = 2 + i
    style_cell(ws3, row, c, WHITE_BOLD, HEADER_BG)
    ws3.cell(row=row, column=c).value = h

# Years: 2025 to 2150 in 5-year steps
years = list(range(2025, 2155, 5))
TIMELINE_START_ROW = 7

# References to Outcomes sheet
# Arrival years: Outcomes!L8, L9, L10, L11, L12
arrival_refs = [f"Outcomes!L{OUTCOME_START_ROW + i}" for i in range(5)]
k_ref = f"Inputs!{K_CELL}"

# MoM weight refs: Outcomes!K15..K19
mom_weight_refs = [f"Outcomes!K{MOM_START_ROW + i}" for i in range(5)]

for i, yr in enumerate(years):
    r = TIMELINE_START_ROW + i

    # Year column
    cell = style_cell(ws3, r, 2, WHITE_BOLD)
    cell.value = yr
    cell.number_format = "0"

    # 5 outcome columns (C through G)
    for j in range(5):
        col = 3 + j
        # =ROUND(100/(1+EXP(-k*(year-arrival))), 0)
        formula = f"=ROUND(100/(1+EXP(-{k_ref}*(B{r}-{arrival_refs[j]}))),0)"
        cell = style_cell(ws3, r, col)
        cell.value = formula
        cell.number_format = "0"

    # Matrix column (I = 9)
    # = C*w1 + D*w2 + E*w3 + F*w4 + G*w5
    parts = []
    outcome_cols = ["C", "D", "E", "F", "G"]
    for j in range(5):
        parts.append(f"{outcome_cols[j]}{r}*{mom_weight_refs[j]}")
    formula = f"=ROUND({'+'.join(parts)},0)"
    cell = style_cell(ws3, r, 9, GREEN_FONT)
    cell.value = formula
    cell.number_format = "0"

TIMELINE_END_ROW = TIMELINE_START_ROW + len(years) - 1

# ── Chart ───────────────────────────────────────────────────

chart = LineChart()
chart.title = "Moment of Matrix Over Time"
chart.y_axis.title = "Score (0-100)"
chart.x_axis.title = "Year"
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 100
chart.width = 25
chart.height = 15
chart.style = 10

# Matrix line (main)
matrix_data = Reference(ws3, min_col=9, min_row=TIMELINE_START_ROW - 1, max_row=TIMELINE_END_ROW)
years_ref = Reference(ws3, min_col=2, min_row=TIMELINE_START_ROW, max_row=TIMELINE_END_ROW)
chart.add_data(matrix_data, titles_from_data=True)
chart.set_categories(years_ref)

# Individual outcome lines
for j in range(5):
    col = 3 + j
    data = Reference(ws3, min_col=col, min_row=TIMELINE_START_ROW - 1, max_row=TIMELINE_END_ROW)
    chart.add_data(data, titles_from_data=True)

# Style the lines
colors = ["34D399", "A78BFA", "F97316", "06B6D4", "EC4899", "EAB308"]
for i, series in enumerate(chart.series):
    series.graphicalProperties.line.width = 25000 if i == 0 else 12000
    series.graphicalProperties.line.solidFill = colors[i]
    if i > 0:
        series.graphicalProperties.line.dashStyle = "dash"

ws3.add_chart(chart, "B30")

# ═══════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════

filepath = "/Users/vasbakos/Documents/personal/matrix/simulator/app/matrix-outcome-model.xlsx"
wb.save(filepath)
print(f"Saved to {filepath}")
